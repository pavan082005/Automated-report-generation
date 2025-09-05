# reports_module.py

# ===============================
# All Imports (moved to top)
# ===============================
import os
import time
import glob
import shutil
import smtplib
import tempfile
import json
import mimetypes
from pathlib import Path
from io import BytesIO
from email.message import EmailMessage

import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.dates import AutoDateLocator, DateFormatter
from PIL import Image as PILImage, ImageEnhance

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
)

# ===============================
# Environment and Paths
# ===============================
load_dotenv()
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")

DATA_DIR = Path("generated_reports")
DATA_DIR.mkdir(exist_ok=True)
STATS_FILE = DATA_DIR / "report_stats.json"

if not STATS_FILE.exists():
    STATS_FILE.write_text("{}")

REPORT_RECIPIENTS = {
    "csv": ["pavan082005@gmail.com"],
    "excel": ["pavan082005@gmail.com"],
    "pdf": ["pavan082005@gmail.com"],
}

# Matplotlib defaults for readable charts (consistent across functions)
matplotlib.rcParams.update(
    {
        "figure.dpi": 150,
        "axes.titlesize": 13,
        "axes.labelsize": 10,
        "xtick.labelsize": 9,
        "ytick.labelsize": 9,
        "legend.fontsize": 9,
        "font.size": 9,
        "axes.grid": True,
        "grid.linestyle": "--",
        "grid.alpha": 0.35,
    }
)

PALETTE = [
    "#2b8aee",
    "#4ecdc4",
    "#FF6B6B",
    "#ffd166",
    "#845ef7",
    "#ff9f1c",
    "#6a4c93",
]


# ===============================
# Helper Functions
# ===============================
def safe_json_load(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def safe_json_write(path, obj):
    tmp = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=4)
    os.replace(tmp, str(path))


def update_report_stats(report_type: str, filepath: str, df: pd.DataFrame):
    stats = safe_json_load(STATS_FILE) or {}

    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    summary = {
        "timestamp": ts,
        "path": str(filepath),
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "numeric_columns": numeric_cols,
        "top_categories": {},
    }

    # Top categories for first 5 object columns
    obj_cols = df.select_dtypes(include="object").columns.tolist()[:5]
    for col in obj_cols:
        try:
            top_vals = df[col].value_counts().nlargest(5).to_dict()
            summary["top_categories"][col] = top_vals
        except Exception:
            summary["top_categories"][col] = {}

    if report_type not in stats:
        stats[report_type] = []
    stats[report_type].append(summary)
    stats[report_type] = stats[report_type][-10:]  # Keep last 10 reports

    # write atomically
    safe_json_write(STATS_FILE, stats)
    return summary


def get_latest_report(report_type: str):
    ext = "xlsx" if report_type == "excel" else report_type
    files = glob.glob(str(DATA_DIR / f"report_*.{ext}"))
    if not files:
        return None
    return max(files, key=os.path.getctime)


def log_report_sent(to_email: str, report_type: str, filepath: str):
    stats = safe_json_load(STATS_FILE) or {}

    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    sent_entry = {"timestamp": ts, "to_email": to_email, "path": str(filepath)}

    if "emails_sent" not in stats:
        stats["emails_sent"] = {}
    if report_type not in stats["emails_sent"]:
        stats["emails_sent"][report_type] = []

    stats["emails_sent"][report_type].append(sent_entry)
    stats["emails_sent"][report_type] = stats["emails_sent"][report_type][-20:]  # last 20

    # Update last_sent info in report stats
    if report_type in stats and stats[report_type]:
        stats[report_type][-1]["last_sent_to"] = to_email
        stats[report_type][-1]["last_sent_at"] = ts

    safe_json_write(STATS_FILE, stats)


# ===============================
# Plot helpers
# ===============================
def fig_to_buffer(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def save_fig(fig, path):
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def annotate_bars(ax, bars, fmt="{:.0f}"):
    for b in bars:
        h = b.get_height()
        if np.isfinite(h):
            ax.annotate(
                fmt.format(h),
                xy=(b.get_x() + b.get_width() / 2, h),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=8,
            )


def create_bar(labels, values, title, xlabel=None, ylabel="Count", rotate=30, width=7.0, height=3.0):
    fig, ax = plt.subplots(figsize=(width, height))
    bars = ax.bar(range(len(values)), values, color=[PALETTE[i % len(PALETTE)] for i in range(len(values))])
    ax.set_xticks(range(len(values)))
    ax.set_xticklabels(labels, rotation=rotate, ha="right")
    ax.set_ylabel(ylabel)
    if xlabel:
        ax.set_xlabel(xlabel)
    ax.set_title(title)
    annotate_bars(ax, bars)
    plt.tight_layout()
    return fig


def create_hist(values, title, xlabel=None, width=7.0, height=3.0):
    fig, ax = plt.subplots(figsize=(width, height))
    vals = np.array(values)
    vals = vals[~np.isnan(vals)]
    bins = min(30, max(6, int(np.sqrt(vals.size)))) if vals.size > 0 else 10
    ax.hist(vals, bins=bins)
    ax.set_title(title)
    if xlabel:
        ax.set_xlabel(xlabel)
    ax.set_ylabel("Frequency")
    plt.tight_layout()
    return fig


def create_boxplot(arrays, labels, title, width=7.0, height=3.0):
    fig, ax = plt.subplots(figsize=(width, height))
    ax.boxplot(arrays, labels=labels, patch_artist=True)
    ax.set_title(title)
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right")
    plt.tight_layout()
    return fig


def create_heatmap(corr_df, title, width=7.0, height=6.0):
    fig, ax = plt.subplots(figsize=(width, height))
    im = ax.imshow(corr_df.values, cmap="RdBu_r", vmin=-1, vmax=1)
    ax.set_xticks(range(len(corr_df.columns)))
    ax.set_xticklabels(corr_df.columns, rotation=45, ha="right")
    ax.set_yticks(range(len(corr_df.index)))
    ax.set_yticklabels(corr_df.index)
    ax.set_title(title)
    # annotate values
    for (i, j), val in np.ndenumerate(corr_df.values):
        ax.text(j, i, f"{val:.2f}", ha="center", va="center", fontsize=7)
    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    plt.tight_layout()
    return fig


# ===============================
# CSV generation + visual summary
# ===============================
def generate_csv(uploaded_file):
    """
    Saves uploaded CSV to generated_reports and creates visualizations + a small summary PDF.
    Returns path to saved CSV (string). Visual assets and summary PDF are written next to CSV.
    uploaded_file can be a file-like with .file or a string path.
    """
    ts = time.strftime("%Y%m%d_%H%M%S")
    # Read CSV from file-like object (uploaded_file.file) or path
    if hasattr(uploaded_file, "file"):
        df = pd.read_csv(uploaded_file.file)
    else:
        df = pd.read_csv(uploaded_file)

    output_file = DATA_DIR / f"report_{ts}.csv"
    df.to_csv(output_file, index=False)

    # Create an assets folder for this report
    assets_dir = DATA_DIR / f"report_{ts}_assets"
    assets_dir.mkdir(parents=True, exist_ok=True)

    # Basic detection
    numeric_df = df.apply(pd.to_numeric, errors="coerce")
    numeric_cols = numeric_df.select_dtypes(include=[np.number]).columns.tolist()
    cat_cols = [c for c in df.columns if (df[c].dtype == object or str(df[c].dtype).startswith("category"))]

    # detect date-like column
    def detect_date_column(df_):
        priority = ["timestamp", "time", "date", "created_at", "datetime"]
        for name in priority:
            for c in df_.columns:
                if name.lower() in c.lower():
                    try:
                        parsed = pd.to_datetime(df_[c], errors="coerce")
                        if parsed.notna().sum() > 0:
                            return c
                    except Exception:
                        pass
        best = None
        best_ratio = 0.0
        for c in df_.columns:
            try:
                parsed = pd.to_datetime(df_[c], errors="coerce")
            except Exception:
                parsed = pd.Series([pd.NaT] * len(df_))
            ratio = parsed.notna().mean()
            if ratio > best_ratio and ratio >= 0.25:
                best_ratio = ratio
                best = c
        return best

    date_col = detect_date_column(df)
    if date_col:
        try:
            df["_parsed_date"] = pd.to_datetime(df[date_col], errors="coerce")
            df_dates = df[df["_parsed_date"].notna()].copy()
        except Exception:
            df_dates = pd.DataFrame()
    else:
        df_dates = pd.DataFrame()

    created_images = []

    # 1) Histograms + boxplots for top numeric columns (up to 3)
    if numeric_cols:
        num_for_hist = numeric_cols[:3]
        for col in num_for_hist:
            try:
                hist_fig = create_hist(numeric_df[col].values, title=f"Distribution of '{col}'", xlabel=col)
                hist_path = assets_dir / f"hist_{col}.png"
                save_fig(hist_fig, hist_path)
                created_images.append(str(hist_path))
            except Exception:
                pass

        # boxplots (top 4 by std)
        try:
            top_k = sorted(numeric_cols, key=lambda c: numeric_df[c].std(skipna=True) if numeric_df[c].notna().any() else -1, reverse=True)[:4]
            arrays = [numeric_df[c].dropna().values for c in top_k]
            if any(len(a) > 0 for a in arrays):
                box_fig = create_boxplot(arrays, top_k, title="Boxplots (top variable numeric columns)")
                box_path = assets_dir / "boxplots.png"
                save_fig(box_fig, box_path)
                created_images.append(str(box_path))
        except Exception:
            pass

    # 2) Correlation heatmap if >=2 numeric
    if len(numeric_cols) >= 2:
        try:
            corr = numeric_df[numeric_cols].corr().round(2)
            heat_fig = create_heatmap(corr, title="Correlation matrix (pearson)")
            heat_path = assets_dir / "correlation_heatmap.png"
            save_fig(heat_fig, heat_path)
            created_images.append(str(heat_path))
        except Exception:
            pass

    # 3) Categorical bar + donut for best categorical column
    chosen_cat = None
    candidates = [c for c in cat_cols if 1 < df[c].nunique(dropna=True) <= 200]
    if candidates:
        chosen_cat = max(candidates, key=lambda c: df[c].notna().sum())
    else:
        fallback = sorted(df.columns, key=lambda c: df[c].nunique(dropna=True))
        for c in fallback:
            if df[c].nunique(dropna=True) > 1 and df[c].nunique(dropna=True) <= 500:
                chosen_cat = c
                break

    if chosen_cat:
        try:
            counts = df[chosen_cat].value_counts().nlargest(12)
            bar_fig = create_bar(list(counts.index.astype(str)), counts.values, title=f"Top values in '{chosen_cat}'", xlabel=chosen_cat, ylabel="Count", rotate=35)
            bar_path = assets_dir / f"bar_{chosen_cat}.png"
            save_fig(bar_fig, bar_path)
            created_images.append(str(bar_path))

            # donut
            def create_donut(labels, values, title, width=6.2, height=6.2):
                fig, ax = plt.subplots(figsize=(width, height))
                colors_list = [PALETTE[i % len(PALETTE)] for i in range(len(values))]
                def autopct_gen(vals):
                    def autopct(pct):
                        total = np.sum(vals)
                        val = int(round(pct * total / 100.0))
                        return f"{pct:.1f}%\n({val})" if pct >= 2 else ""
                    return autopct
                wedges, texts, autotexts = ax.pie(values, labels=None, startangle=140, colors=colors_list,
                                                  wedgeprops=dict(width=0.36, edgecolor='white', linewidth=0.8),
                                                  autopct=autopct_gen(values), pctdistance=0.78)
                ax.axis('equal')
                ax.legend(wedges, labels, title="Categories", loc='center left', bbox_to_anchor=(1.05, 0.5), frameon=False)
                ax.set_title(title, fontsize=12)
                plt.tight_layout()
                return fig

            donut_fig = create_donut(list(counts.index.astype(str)), counts.values, title=f"Share in '{chosen_cat}'")
            donut_path = assets_dir / f"donut_{chosen_cat}.png"
            save_fig(donut_fig, donut_path)
            created_images.append(str(donut_path))
        except Exception:
            pass

    # 4) Time-series if detected
    if date_col and not df_dates.empty and len(numeric_cols) > 0:
        try:
            best_num = max(numeric_cols, key=lambda c: numeric_df[c].notna().sum())
            ts = df_dates.set_index("_parsed_date").resample("D")[best_num].sum().fillna(0)
            if ts.shape[0] >= 2:
                fig, ax = plt.subplots(figsize=(8, 3.2))
                ax.plot(ts.index, ts.values, marker="o", linewidth=1.6)
                ax.fill_between(ts.index, ts.values, alpha=0.18)
                ax.set_title(f"Daily aggregated {best_num}")
                locator = AutoDateLocator()
                formatter = DateFormatter("%Y-%m-%d")
                ax.xaxis.set_major_locator(locator)
                ax.xaxis.set_major_formatter(formatter)
                plt.setp(ax.get_xticklabels(), rotation=30, ha="right")
                plt.tight_layout()
                ts_path = assets_dir / f"timeseries_{best_num}.png"
                save_fig(fig, ts_path)
                created_images.append(str(ts_path))
        except Exception:
            pass

    # 5) Create a small PDF summary embedding created_images (if any)
    summary_pdf_path = DATA_DIR / f"report_{ts}_summary.pdf"
    try:
        if created_images:
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
                base_font = "DejaVuSans"
            except Exception:
                base_font = "Helvetica"

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("Title", parent=styles["Title"], fontName=base_font, fontSize=18, alignment=1)
            normal_style = ParagraphStyle("Normal", parent=styles["Normal"], fontName=base_font, fontSize=9)

            doc = SimpleDocTemplate(str(summary_pdf_path), pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
            elems = []
            elems.append(Paragraph("CSV Report - Visual Summary", title_style))
            elems.append(Spacer(1, 8))
            elems.append(Paragraph(f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
            elems.append(Spacer(1, 12))

            # preview table (first 6 rows)
            elems.append(Paragraph("Data preview (first 6 rows)", normal_style))
            preview = df.head(6)
            table_data = [list(preview.columns)]
            for _, r in preview.iterrows():
                table_data.append([("" if pd.isna(v) else str(v)) for v in r])
            preview_table = Table(table_data, repeatRows=1)
            preview_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                        ("FONTNAME", (0, 0), (-1, -1), base_font),
                    ]
                )
            )
            elems.append(preview_table)
            elems.append(Spacer(1, 10))

            # Add images (max 3 per page)
            for i, img_path in enumerate(created_images):
                if i and i % 3 == 0:
                    elems.append(PageBreak())
                try:
                    im = PILImage.open(img_path)
                    iw, ih = im.size
                    max_w = 6.5 * inch
                    scale = min(1.0, max_w / iw)
                    display_h = ih * (max_w / iw)
                    elems.append(Image(img_path, width=max_w, height=display_h))
                    elems.append(Spacer(1, 8))
                except Exception:
                    pass

            doc.build(elems)
    except Exception:
        pass

    # Update stats
    try:
        update_report_stats("csv", output_file, df)
    except Exception:
        pass

    return str(output_file)


# ===============================
# Excel generation (full)
# ===============================
def generate_excel(
    uploaded_file,
    data_dir=None,
    max_numeric_charts=10,
    max_cat_bars=10,
    figsize=(6, 4),
    include_scatter_matrix=False,
):
    """
    Reads uploaded CSV, writes an Excel with multiple sheets (Raw Data, Summary, Data Dictionary, Missing Values, Top Categories),
    creates charts as PNGs and embeds them into a 'Charts' sheet.
    Returns path to generated .xlsx file.
    """
    if data_dir:
        base_dir = Path(data_dir)
    else:
        base_dir = DATA_DIR
    base_dir.mkdir(parents=True, exist_ok=True)

    ts = time.strftime("%Y%m%d_%H%M%S")
    # Read CSV
    if hasattr(uploaded_file, "file"):
        df = pd.read_csv(uploaded_file.file)
    else:
        df = pd.read_csv(uploaded_file)

    output_file = base_dir / f"report_{ts}.xlsx"

    numeric = df.select_dtypes(include="number")
    categorical = df.select_dtypes(exclude="number")

    describe = df.describe(include="all").transpose()
    if not numeric.empty:
        more_stats = pd.DataFrame({"skew": numeric.skew(), "kurtosis": numeric.kurtosis()})
        describe = describe.join(more_stats, how="outer")

    missing_count = df.isnull().sum()
    missing_pct = (missing_count / len(df)) * 100
    missing_df = pd.concat([missing_count, missing_pct], axis=1)
    missing_df.columns = ["missing_count", "missing_pct"]

    data_dict = []
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(3).tolist()
        data_dict.append(
            {
                "column": col,
                "dtype": str(df[col].dtype),
                "n_unique": df[col].nunique(dropna=True),
                "n_missing": int(df[col].isnull().sum()),
                "sample_values": ", ".join(sample),
            }
        )
    data_dict_df = pd.DataFrame(data_dict).set_index("column")

    top_cats = {}
    for col in categorical.columns:
        vc = df[col].value_counts(dropna=False).head(max_cat_bars)
        top_cats[col] = vc

    # Write base sheets
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        describe.to_excel(writer, sheet_name="Summary")
        data_dict_df.to_excel(writer, sheet_name="Data Dictionary")
        missing_df.to_excel(writer, sheet_name="Missing Values")
        top_cats_frames = []
        for col, vc in top_cats.items():
            tmp = vc.reset_index()
            tmp.columns = [f"{col}_value", f"{col}_count"]
            top_cats_frames.append(tmp)
        if top_cats_frames:
            max_rows = max(len(t) for t in top_cats_frames)
            expanded = []
            for t in top_cats_frames:
                t2 = t.reindex(range(max_rows))
                expanded.append(t2.reset_index(drop=True))
            merged = pd.concat(expanded, axis=1)
            merged.to_excel(writer, sheet_name="Top Categories", index=False)
        else:
            pd.DataFrame().to_excel(writer, sheet_name="Top Categories")

    # Create charts and save in temporary dir
    tmpdir = Path(tempfile.mkdtemp(prefix="excel_report_"))
    created_images = []
    try:
        num_cols = numeric.columns.tolist()[:max_numeric_charts]
        for col in num_cols:
            try:
                plt.figure(figsize=figsize)
                df[col].dropna().plot.hist(bins=30)
                plt.title(f"Histogram: {col}")
                plt.xlabel(col)
                hist_path = tmpdir / f"hist_{col}.png"
                plt.tight_layout()
                plt.savefig(hist_path)
                plt.close()
                created_images.append((hist_path, f"Histogram - {col}"))

                plt.figure(figsize=figsize)
                df[col].dropna().plot.box()
                plt.title(f"Boxplot: {col}")
                plt.ylabel(col)
                box_path = tmpdir / f"box_{col}.png"
                plt.tight_layout()
                plt.savefig(box_path)
                plt.close()
                created_images.append((box_path, f"Boxplot - {col}"))
            except Exception:
                continue

        if numeric.shape[1] >= 2:
            try:
                corr = numeric.corr()
                plt.figure(figsize=(max(6, corr.shape[0] * 0.6), max(4, corr.shape[1] * 0.5)))
                im = plt.imshow(corr, interpolation="nearest", aspect="auto", cmap="RdBu_r", vmin=-1, vmax=1)
                plt.colorbar(im)
                plt.xticks(range(len(corr.columns)), corr.columns, rotation=90)
                plt.yticks(range(len(corr.index)), corr.index)
                plt.title("Correlation matrix")
                heat_path = tmpdir / "correlation_heatmap.png"
                plt.tight_layout()
                plt.savefig(heat_path)
                plt.close()
                created_images.append((heat_path, "Correlation heatmap"))
            except Exception:
                pass

        for col, vc in top_cats.items():
            try:
                plt.figure(figsize=figsize)
                labels = [str(x) for x in vc.index]
                heights = vc.values
                plt.bar(range(len(heights)), heights)
                plt.xticks(range(len(heights)), labels, rotation=45, ha="right")
                plt.title(f"Top {len(heights)} values: {col}")
                plt.tight_layout()
                bar_path = tmpdir / f"bar_{col}.png"
                plt.savefig(bar_path)
                plt.close()
                created_images.append((bar_path, f"Top values - {col}"))
            except Exception:
                pass

        # timeseries detection (look for date-like column)
        date_col = None
        priority = ["timestamp", "time", "date", "created_at", "datetime"]
        for name in priority:
            for c in df.columns:
                if name.lower() in c.lower():
                    try:
                        parsed = pd.to_datetime(df[c], errors="coerce")
                        if parsed.notna().sum() > 0:
                            date_col = c
                            df["_parsed_date"] = parsed
                            break
                    except Exception:
                        pass
            if date_col:
                break

        if not date_col:
            # fallback: heuristics
            best = None
            best_ratio = 0.0
            for c in df.columns:
                try:
                    parsed = pd.to_datetime(df[c], errors="coerce")
                except Exception:
                    parsed = pd.Series([pd.NaT] * len(df))
                ratio = parsed.notna().mean()
                if ratio > best_ratio and ratio >= 0.25:
                    best_ratio = ratio
                    best = c
            if best:
                date_col = best
                df["_parsed_date"] = pd.to_datetime(df[best], errors="coerce")

        if date_col and not numeric.empty:
            try:
                dcol = "_parsed_date"
                ncol = numeric.columns[0]
                ts_df = df[[dcol, ncol]].dropna().sort_values(by=dcol)
                if not ts_df.empty:
                    plt.figure(figsize=(8, 4))
                    plt.plot(ts_df[dcol], ts_df[ncol], marker="o", linestyle="-")
                    plt.title(f"Timeseries: {ncol} over {dcol}")
                    plt.xlabel(dcol)
                    plt.ylabel(ncol)
                    ts_path = tmpdir / "timeseries.png"
                    plt.tight_layout()
                    plt.savefig(ts_path)
                    plt.close()
                    created_images.append((ts_path, "Timeseries"))
            except Exception:
                pass

        # scatter matrix if requested
        if include_scatter_matrix and numeric.shape[1] >= 2:
            try:
                cols_for_matrix = numeric.columns.tolist()[:6]
                pd.plotting.scatter_matrix(numeric[cols_for_matrix].dropna(), figsize=(8, 8))
                sm_path = tmpdir / "scatter_matrix.png"
                plt.tight_layout()
                plt.savefig(sm_path)
                plt.close()
                created_images.append((sm_path, "Scatter matrix"))
            except Exception:
                pass

        # Now embed images into the Excel file (Charts sheet)
        wb = load_workbook(output_file)
        if "Charts" in wb.sheetnames:
            ws = wb["Charts"]
        else:
            ws = wb.create_sheet("Charts")

        row = 1
        for img_path, title in created_images:
            try:
                ws.cell(row=row, column=1, value=title)
                row += 1
                img = XLImage(str(img_path))
                anchor_cell = f"A{row}"
                ws.add_image(img, anchor_cell)
                row += 20
            except Exception:
                continue

        wb.save(output_file)
    finally:
        # cleanup tmpdir
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass

    update_report_stats("excel", output_file, df)
    return str(output_file)


# ===============================
# PDF generation (full)
# ===============================
def generate_pdf(
    uploaded_file,
    output_dir=None,
    logo_path=None,
    title="AUTOMATIC DATA REPORT",
    watermark_opacity=0.06,
    max_cat_display=12,
    dpi=150,
):
    """
    Reads CSV and produces a multipage PDF report with charts and textual summaries.
    Returns path to generated PDF.
    """
    ts = time.strftime("%Y%m%d_%H%M%S")

    if output_dir:
        outdir = Path(output_dir)
    else:
        outdir = DATA_DIR
    outdir.mkdir(parents=True, exist_ok=True)
    output_file = outdir / f"report_{ts}.pdf"

    if hasattr(uploaded_file, "file"):
        df = pd.read_csv(uploaded_file.file)
    else:
        df = pd.read_csv(uploaded_file)

    df.columns = [str(c) for c in df.columns]

    numeric_df = df.apply(pd.to_numeric, errors="coerce")
    numeric_cols = numeric_df.select_dtypes(include=[np.number]).columns.tolist()
    cat_cols = [c for c in df.columns if (df[c].dtype == object or str(df[c].dtype).startswith("category"))]

    # detect date column
    def detect_date_column(df_):
        priority = ["timestamp", "time", "date", "created_at", "datetime"]
        for name in priority:
            for c in df_.columns:
                if name.lower() in c.lower():
                    try:
                        parsed = pd.to_datetime(df_[c], errors="coerce")
                        if parsed.notna().sum() > 0:
                            return c
                    except Exception:
                        pass
        best = None
        best_ratio = 0.0
        for c in df_.columns:
            try:
                parsed = pd.to_datetime(df_[c], errors="coerce")
            except Exception:
                parsed = pd.Series([pd.NaT] * len(df_))
            ratio = parsed.notna().mean()
            if ratio > best_ratio and ratio >= 0.25:
                best_ratio = ratio
                best = c
        return best

    date_col = detect_date_column(df)
    if date_col:
        df["_parsed_date"] = pd.to_datetime(df[date_col], errors="coerce")
        df_dates = df[df["_parsed_date"].notna()].copy()
    else:
        df_dates = pd.DataFrame()

    matplotlib.rcParams.update({"figure.dpi": dpi})

    # helper chart creators (reuse earlier helpers)
    def fig_to_buf_local(fig):
        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    def create_donut(labels, values, title, width=6.2, height=6.2):
        fig, ax = plt.subplots(figsize=(width, height))
        colors_list = [PALETTE[i % len(PALETTE)] for i in range(len(values))]
        def autopct_gen(vals):
            def autopct(pct):
                total = np.sum(vals)
                val = int(round(pct * total / 100.0))
                return f"{pct:.1f}%\n({val})" if pct >= 2 else ""
            return autopct
        wedges, texts, autotexts = ax.pie(values, labels=None, startangle=140, colors=colors_list,
                                          wedgeprops=dict(width=0.36, edgecolor='white', linewidth=0.8),
                                          autopct=autopct_gen(values), pctdistance=0.78)
        ax.axis('equal')
        ax.legend(wedges, labels, title="Categories", loc='center left', bbox_to_anchor=(1.05, 0.5), frameon=False)
        ax.set_title(title, fontsize=14)
        plt.tight_layout()
        return fig

    # register font if available
    try:
        pdfmetrics.registerFont(TTFont("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
        base_font = "DejaVuSans"
    except Exception:
        base_font = "Helvetica"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontName=base_font, fontSize=20, leading=24, alignment=1)
    date_style = ParagraphStyle("Date", parent=styles["Normal"], fontName=base_font, fontSize=9, alignment=1)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontName=base_font, fontSize=12)
    normal_style = ParagraphStyle("Normal", parent=styles["Normal"], fontName=base_font, fontSize=10)

    doc = SimpleDocTemplate(str(output_file), pagesize=letter, rightMargin=36, leftMargin=36, topMargin=48, bottomMargin=36)
    elements = []

    # watermark preparation
    watermark_reader = None
    if logo_path and os.path.exists(logo_path):
        try:
            pil = PILImage.open(logo_path).convert("RGBA")
            page_w, page_h = letter
            target_w = int(page_w * 0.6)
            ratio = target_w / max(1, pil.width)
            target_h = int(pil.height * ratio)
            pil = pil.resize((target_w, target_h), PILImage.LANCZOS)
            alpha = pil.split()[3]
            alpha = ImageEnhance.Brightness(alpha).enhance(max(0.0, min(1.0, watermark_opacity)))
            pil.putalpha(alpha)
            pil = pil.rotate(12, expand=True)
            bio = BytesIO()
            pil.save(bio, format="PNG")
            bio.seek(0)
            watermark_reader = ImageReader(bio)
        except Exception:
            watermark_reader = None

    def draw_background(canvas, doc_obj):
        canvas.saveState()
        canvas.setFillColor(colors.white)
        canvas.rect(0, 0, doc_obj.pagesize[0], doc_obj.pagesize[1], stroke=0, fill=1)
        if watermark_reader is not None:
            try:
                img_w, img_h = watermark_reader.getSize()
                x = (doc_obj.pagesize[0] - img_w) / 2
                y = (doc_obj.pagesize[1] - img_h) / 2
                canvas.drawImage(watermark_reader, x, y, width=img_w, height=img_h, mask='auto')
            except Exception:
                pass
        if logo_path and os.path.exists(logo_path):
            try:
                w, h = 52, 52
                x = doc_obj.pagesize[0] - (w + 34)
                y = doc_obj.pagesize[1] - (h + 34)
                canvas.drawImage(logo_path, x, y, width=w, height=h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        canvas.setFont(base_font, 8)
        canvas.setFillColor(colors.black)
        canvas.drawRightString(doc_obj.pagesize[0] - 30, 20, f"Page {doc_obj.page}")
        canvas.restoreState()

    # Header
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(time.strftime("%d %B %Y"), date_style))
    elements.append(Spacer(1, 12))

    # Preview table
    elements.append(Paragraph("Data preview (first 8 rows)", heading_style))
    preview = df.head(8)
    table_data = [list(preview.columns)]
    for _, r in preview.iterrows():
        table_data.append([("" if pd.isna(v) else str(v)) for v in r])
    preview_table = Table(table_data, repeatRows=1)
    preview_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")), ("FONTNAME", (0, 0), (-1, -1), base_font)]))
    elements.append(preview_table)
    elements.append(Spacer(1, 12))

    # Numeric summary
    elements.append(Paragraph("Numeric summary (basic stats)", heading_style))
    if numeric_cols:
        stats_df = numeric_df[numeric_cols].describe().T.round(3)
        stats_table_data = [["column"] + stats_df.columns.tolist()]
        for col in stats_df.index:
            stats_table_data.append([col] + [str(stats_df.loc[col, c]) for c in stats_df.columns])
        st = Table(stats_table_data, repeatRows=1)
        st.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")), ("FONTNAME", (0, 0), (-1, -1), base_font)]))
        elements.append(st)
    else:
        elements.append(Paragraph("No numeric columns detected.", normal_style))
    elements.append(PageBreak())

    # Categorical charts
    elements.append(Paragraph("Top categorical distributions", heading_style))
    chosen_cat = None
    candidates = [c for c in cat_cols if 1 < df[c].nunique(dropna=True) <= 200]
    if candidates:
        chosen_cat = max(candidates, key=lambda c: df[c].notna().sum())
    else:
        fallback = sorted(df.columns, key=lambda c: df[c].nunique(dropna=True))
        for c in fallback:
            if df[c].nunique(dropna=True) > 1 and df[c].nunique(dropna=True) <= 500:
                chosen_cat = c
                break

    chart_images = []
    if chosen_cat:
        counts = df[chosen_cat].value_counts().nlargest(max_cat_display)
        bar_fig = create_bar(list(counts.index.astype(str)), counts.values, title=f"Top values in '{chosen_cat}'", xlabel=chosen_cat, ylabel="Count", rotate=35)
        chart_images.append(fig_to_buf_local(bar_fig))
        donut_fig = create_donut(list(counts.index.astype(str)), counts.values, title=f"Share in '{chosen_cat}'")
        chart_images.append(fig_to_buf_local(donut_fig))

        for buf in chart_images:
            elements.append(Image(buf, width=6.8 * inch, height=3.2 * inch))
            elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"Figure: Distribution and share of top values in <b>{chosen_cat}</b>.", normal_style))
    else:
        elements.append(Paragraph("No suitable categorical column found for bar/donut charts.", normal_style))
    elements.append(PageBreak())

    # Time-series
    elements.append(Paragraph("Time-series & trend analysis", heading_style))
    if date_col and not df_dates.empty and numeric_cols:
        best_num = max(numeric_cols, key=lambda c: numeric_df[c].notna().sum())
        ts = df_dates.set_index("_parsed_date").resample("D")[best_num].sum().fillna(0)
        if ts.shape[0] >= 3:
            line_fig = create_bar([str(x.date()) for x in ts.index], ts.values, title=f"Daily aggregated {best_num}", xlabel=date_col, ylabel=best_num, rotate=45, width=8.0, height=3.2)
            elements.append(Image(fig_to_buf_local(line_fig), width=6.8 * inch, height=3.2 * inch))
            elements.append(Spacer(1, 8))
            try:
                hours = df_dates["_parsed_date"].dt.hour.dropna().astype(int)
                hourly_counts = hours.value_counts().reindex(range(24), fill_value=0).sort_index()
                bar_hour = create_bar([f"{h:02d}" for h in range(24)], hourly_counts.values, title="Activity by hour (0-23)", xlabel="Hour", ylabel="Count", rotate=0, width=6.8, height=2.4)
                elements.append(Image(fig_to_buf_local(bar_hour), width=6.8 * inch, height=2.2 * inch))
                elements.append(Paragraph("Figure: Daily aggregate of the top numeric column and hourly activity distribution.", normal_style))
            except Exception:
                pass
        else:
            elements.append(Paragraph("Not enough time-series points to build a day-level trend (need >=3 days).", normal_style))
    else:
        elements.append(Paragraph("No date + numeric combination found for time-series analysis.", normal_style))
    elements.append(PageBreak())

    # Distributions & correlations
    elements.append(Paragraph("Distributions & outliers", heading_style))
    if numeric_cols:
        hist_col = max(numeric_cols, key=lambda c: numeric_df[c].var(skipna=True) if numeric_df[c].notna().any() else -1)
        hist_fig = create_hist(numeric_df[hist_col].values, title=f"Distribution of '{hist_col}'", xlabel=hist_col)
        elements.append(Image(fig_to_buf_local(hist_fig), width=6.8 * inch, height=3.2 * inch))
        elements.append(Spacer(1, 8))

        top_k = sorted(numeric_cols, key=lambda c: numeric_df[c].std(skipna=True) if numeric_df[c].notna().any() else -1, reverse=True)[:4]
        arrays = [numeric_df[c].dropna().values for c in top_k]
        if any(len(a) > 0 for a in arrays):
            bp = create_boxplot(arrays, top_k, title="Boxplots (top variable numeric columns)")
            elements.append(Image(fig_to_buf_local(bp), width=6.8 * inch, height=3.2 * inch))
            elements.append(Paragraph("Figure: Histogram and boxplots for key numeric columns.", normal_style))
    else:
        elements.append(Paragraph("No numeric columns detected for distribution plots.", normal_style))
    elements.append(PageBreak())

    # Correlation & scatter
    elements.append(Paragraph("Correlation & relationships", heading_style))
    if len(numeric_cols) >= 2:
        corr = numeric_df[numeric_cols].corr().round(2)
        heat = create_heatmap(corr, title="Correlation matrix (pearson)")
        elements.append(Image(fig_to_buf_local(heat), width=6.8 * inch, height=6.0 * inch))
        abs_corr = corr.abs().where(~np.eye(len(corr), dtype=bool)).stack()
        if not abs_corr.empty:
            top_pair = abs_corr.idxmax()
            c1, c2 = top_pair
            scatter_fig = create_bar([0], [0], title="")  # placeholder; create a simple scatter
            try:
                scatter_fig, ax = plt.subplots(figsize=(6.8, 4.0))
                ax.scatter(numeric_df[c1].dropna(), numeric_df[c2].dropna(), alpha=0.7, s=18)
                ax.set_xlabel(c1)
                ax.set_ylabel(c2)
                ax.set_title(f"Scatter: {c1} vs {c2}")
                plt.tight_layout()
                elements.append(Image(fig_to_buf_local(scatter_fig), width=6.8 * inch, height=4.0 * inch))
                elements.append(Paragraph(f"Figure: Top correlated pair by absolute Pearson correlation: <b>{c1}</b> & <b>{c2}</b>.", normal_style))
            except Exception:
                pass
    else:
        elements.append(Paragraph("Not enough numeric columns for correlation analysis.", normal_style))
    elements.append(PageBreak())

    # Summaries
    elements.append(Paragraph("Summaries (percentiles & active time ranges)", heading_style))
    if date_col and not df_dates.empty:
        try:
            hours = df_dates["_parsed_date"].dt.hour.dropna().astype(int)
            hourly = hours.value_counts().reindex(range(24), fill_value=0).sort_index().values.astype(float)
            # show hourly bar
            bar_hour = create_bar([f"{h:02d}" for h in range(24)], hourly, title="Activity by hour (0-23)", xlabel="Hour", ylabel="Count", rotate=0, width=6.8, height=2.4)
            elements.append(Image(fig_to_buf_local(bar_hour), width=6.8 * inch, height=2.2 * inch))
            elements.append(Paragraph("Figure: Active hours distribution.", normal_style))
        except Exception:
            pass
    elif numeric_cols:
        col = numeric_cols[0]
        vals = numeric_df[col].dropna().values
        if vals.size > 0:
            p75 = int((vals > np.percentile(vals, 75)).sum() / vals.size * 100)
            p50 = int((vals > np.percentile(vals, 50)).sum() / vals.size * 100)
            p25 = int((vals > np.percentile(vals, 25)).sum() / vals.size * 100)
            for label, pct in {">%tile 75": p75, ">%tile 50": p50, ">%tile 25": p25}.items():
                # simple horizontal bar as gauge
                fig, ax = plt.subplots(figsize=(2.6, 0.55))
                ax.barh([0], [pct], height=0.5, color=PALETTE[0])
                ax.set_xlim(0, 100)
                ax.set_yticks([])
                ax.set_xlabel(f"{label}: {int(round(pct))}%")
                for spine in ax.spines.values():
                    spine.set_visible(False)
                plt.tight_layout()
                elements.append(Image(fig_to_buf_local(fig), width=2.6 * inch, height=0.55 * inch))
                elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"Figure: Percent of '{col}' values above selected percentiles.", normal_style))
        else:
            elements.append(Paragraph("Not enough numeric data for percentile summary.", normal_style))
    else:
        elements.append(Paragraph("Not enough data to build summaries.", normal_style))

    elements.append(PageBreak())

    # Automated summary
    elements.append(Paragraph("Automated summary & insights", heading_style))
    insights = []
    if numeric_cols:
        try:
            top_sum_col = max(numeric_cols, key=lambda c: numeric_df[c].sum(skipna=True))
            top_mean_col = max(numeric_cols, key=lambda c: numeric_df[c].mean(skipna=True) if numeric_df[c].notna().any() else -np.inf)
            insights.append(f"- Column with largest total: <b>{top_sum_col}</b>")
            insights.append(f"- Column with highest mean: <b>{top_mean_col}</b>")
        except Exception:
            pass
    if chosen_cat:
        try:
            top_val = df[chosen_cat].value_counts().idxmax()
            insights.append(f"- Most frequent value in <b>{chosen_cat}</b>: {top_val}")
        except Exception:
            pass
    if date_col:
        insights.append(f"- Detected date column: <b>{date_col}</b>")
    if not insights:
        insights.append("- No specific insights detected.")
    elements.append(Paragraph("<br/>".join(insights), normal_style))
    elements.append(Spacer(1, 12))

    # Build PDF
    doc.build(elements, onFirstPage=draw_background, onLaterPages=draw_background)

    try:
        update_report_stats("pdf", output_file, df)
    except Exception:
        pass

    return str(output_file)


# ===============================
# Email functions
# ===============================
def send_email_with_report(to_email: str, report_type: str) -> bool:
    try:
        filepath = get_latest_report(report_type)
        if not filepath:
            print("No report file found.")
            return False

        filename = os.path.basename(filepath)
        msg = EmailMessage()
        msg["Subject"] = f"Your {report_type.upper()} Report"
        msg["From"] = SMTP_USER
        msg["To"] = to_email
        msg.set_content("Please find the attached report.")

        ctype, encoding = mimetypes.guess_type(filename)
        maintype, subtype = ("application", "octet-stream")
        if ctype:
            maintype, subtype = ctype.split("/", 1)

        with open(filepath, "rb") as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=filename)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)

        log_report_sent(to_email, report_type, filepath)
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


def send_sample_message(to_email: str) -> bool:
    try:
        msg = EmailMessage()
        msg["Subject"] = "Sample Report"
        msg["From"] = SMTP_USER
        msg["To"] = to_email
        msg.set_content("This is a sample report.")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)

        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


# ===============================
# Usage example (commented)
# ===============================
# if __name__ == "__main__":
#     # Example: simulate an uploaded file object with attribute .file
#     class U:
#         def __init__(self, path):
#             self.file = open(path, "rb")
#
#     uploaded = U("sample.csv")
#     csv_path = generate_csv(uploaded)
#     print("CSV created:", csv_path)
#     uploaded.file.close()

# ===============================
# Requirements (install these)
# ===============================
# pip install pandas numpy matplotlib pillow python-dotenv openpyxl reportlab
